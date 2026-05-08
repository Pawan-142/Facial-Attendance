"""
firebase_db.py — Firestore Database Layer
Drop-in replacement for models.py using Firebase Firestore.
All function signatures and return formats are identical to models.py.
"""
import csv
import os
from datetime import datetime, date, timedelta
from firebase_client import get_db
from config import EXPORT_DIR


# ── Helpers ────────────────────────────────────────────────────────────────────
def _col(name):
    return get_db().collection(name)


def _doc_to_dict(doc):
    if doc.exists:
        d = doc.to_dict()
        d["_id"] = doc.id
        return d
    return None


# ══════════════════════════════════════════════════════════════
# INIT
# ══════════════════════════════════════════════════════════════
def init_db():
    """Verify Firestore connection (collections auto-create on first write)."""
    try:
        get_db()
        print("[Firebase] Firestore ready")
    except Exception as e:
        raise ConnectionError(f"[Firebase] Connection failed: {e}")


# ══════════════════════════════════════════════════════════════
# STUDENTS
# ══════════════════════════════════════════════════════════════
def add_student(roll_no, name, email="", department="", year=""):
    ref = _col("students").document(roll_no)
    if ref.get().exists:
        return False
    ref.set({
        "roll_no": roll_no, "name": name, "email": email,
        "department": department, "year": year,
        "enrolled_at": datetime.now().isoformat(),
    })
    return True


def update_student(roll_no, name, email="", department="", year=""):
    _col("students").document(roll_no).update({
        "name": name, "email": email,
        "department": department, "year": year,
    })


def get_all_students():
    docs = _col("students").order_by("roll_no").stream()
    return [d.to_dict() for d in docs]


def get_student(roll_no):
    doc = _col("students").document(roll_no).get()
    return doc.to_dict() if doc.exists else None


def delete_student_record(roll_no):
    # Delete all attendance for this student
    for a in _col("attendance").where("roll_no", "==", roll_no).stream():
        a.reference.delete()
    _col("students").document(roll_no).delete()


# ══════════════════════════════════════════════════════════════
# SUBJECTS
# ══════════════════════════════════════════════════════════════
def add_subject(code, name, department="", credits=3):
    # Check for duplicate code
    existing = _col("subjects").where("code", "==", code.upper()).limit(1).stream()
    if any(True for _ in existing):
        return False
    _col("subjects").add({
        "code": code.upper(), "name": name,
        "department": department, "credits": credits,
    })
    return True


def get_all_subjects():
    docs = _col("subjects").order_by("code").stream()
    result = []
    for d in docs:
        row = d.to_dict()
        row["id"] = d.id
        result.append(row)
    return result


def get_subject(subject_id):
    doc = _col("subjects").document(subject_id).get()
    if not doc.exists:
        return None
    row = doc.to_dict()
    row["id"] = doc.id
    return row


def delete_subject(subject_id):
    # Delete sessions and their attendance
    for ses in _col("sessions").where("subject_id", "==", subject_id).stream():
        for a in _col("attendance").where("session_id", "==", ses.id).stream():
            a.reference.delete()
        ses.reference.delete()
    _col("subjects").document(subject_id).delete()


# ══════════════════════════════════════════════════════════════
# SESSIONS
# ══════════════════════════════════════════════════════════════
def create_session(subject_id, faculty="", notes=""):
    today = date.today().isoformat()
    now   = datetime.now().strftime("%H:%M:%S")
    _, ref = _col("sessions").add({
        "subject_id": subject_id,
        "date": today, "start_time": now,
        "faculty": faculty, "notes": notes,
    })
    return ref.id   # string ID


def get_all_sessions():
    sessions = _col("sessions").order_by("date", direction="DESCENDING").stream()
    result = []
    for s in sessions:
        row = s.to_dict()
        row["id"] = s.id
        # Fetch subject info
        sub = _col("subjects").document(row["subject_id"]).get()
        if sub.exists:
            sd = sub.to_dict()
            row["code"] = sd.get("code", "?")
            row["subject_name"] = sd.get("name", "?")
        result.append(row)
    return result


def get_sessions_for_subject(subject_id):
    sessions = (_col("sessions")
                .where("subject_id", "==", subject_id)
                .stream())
    result = []
    sub = _col("subjects").document(subject_id).get()
    sub_data = sub.to_dict() if sub.exists else {}
    for s in sessions:
        row = s.to_dict()
        row["id"] = s.id
        row["code"] = sub_data.get("code", "?")
        row["subject_name"] = sub_data.get("name", "?")
        result.append(row)
    result.sort(key=lambda x: x.get("date", ""), reverse=True)
    return result


def count_sessions(subject_id=None):
    if subject_id:
        docs = _col("sessions").where("subject_id", "==", subject_id).stream()
    else:
        docs = _col("sessions").stream()
    return sum(1 for _ in docs)


# ══════════════════════════════════════════════════════════════
# ATTENDANCE
# ══════════════════════════════════════════════════════════════
def mark_attendance(roll_no, session_id, confidence=None):
    today = date.today().isoformat()
    now   = datetime.now().strftime("%H:%M:%S")
    
    # Fetch subject_id from session to enforce uniqueness per subject per day
    session_doc = _col("sessions").document(session_id).get()
    subject_id  = session_doc.to_dict().get("subject_id") if session_doc.exists else "unknown"
    
    # Composite doc ID enforces uniqueness (roll_no + subject_id + date)
    doc_id = f"{roll_no}__{subject_id}__{today}"
    ref    = _col("attendance").document(doc_id)
    if ref.get().exists:
        return "duplicate"
    try:
        ref.set({
            "roll_no": roll_no, "session_id": session_id,
            "date": today, "time": now, "confidence": confidence,
        })
        return "marked"
    except Exception as e:
        print(f"[Firebase] mark_attendance error: {e}")
        return "error"


def _enrich_attendance(rows):
    """Add student name/dept and subject code/name to raw attendance rows."""
    db = get_db()
    # Batch-fetch unique students using db.get_all
    roll_nos = list({r["roll_no"] for r in rows})
    students = {}
    if roll_nos:
        refs = [db.collection("students").document(rn) for rn in roll_nos]
        docs = db.get_all(refs)
        for doc in docs:
            if doc.exists:
                students[doc.id] = doc.to_dict()

    # Batch-fetch unique sessions → subjects
    session_ids = list({r["session_id"] for r in rows})
    sessions, subjects = {}, {}
    if session_ids:
        s_refs = [db.collection("sessions").document(sid) for sid in session_ids]
        s_docs = db.get_all(s_refs)
        
        subj_ids_to_fetch = set()
        for doc in s_docs:
            if doc.exists:
                sdata = doc.to_dict()
                sessions[doc.id] = sdata
                if sdata.get("subject_id"):
                    subj_ids_to_fetch.add(sdata.get("subject_id"))
        
        if subj_ids_to_fetch:
            sub_refs = [db.collection("subjects").document(sid) for sid in subj_ids_to_fetch]
            sub_docs = db.get_all(sub_refs)
            for doc in sub_docs:
                if doc.exists:
                    subjects[doc.id] = doc.to_dict()

    result = []
    for r in rows:
        rn  = r["roll_no"]
        sid = r["session_id"]
        st  = students.get(rn, {})
        ses = sessions.get(sid, {})
        sub = subjects.get(ses.get("subject_id", ""), {})
        result.append({
            "roll_no":      rn,
            "name":         st.get("name", rn),
            "department":   st.get("department", ""),
            "year":         st.get("year", ""),
            "subject_code": sub.get("code", "?"),
            "subject_name": sub.get("name", "?"),
            "subject_id":   ses.get("subject_id"), # Added for filtering
            "session_id":   sid,                   # Added for completeness
            "date":         r.get("date", ""),
            "time":         r.get("time", ""),
            "confidence":   r.get("confidence"),
        })
    return result


def get_session_attendance(session_id):
    docs = _col("attendance").where("session_id", "==", session_id).stream()
    rows = [d.to_dict() for d in docs]
    rows.sort(key=lambda x: x.get("time", ""))
    return _enrich_attendance(rows)


def get_today_attendance(subject_id=None):
    today = date.today().isoformat()
    q = _col("attendance").where("date", "==", today)
    rows = [d.to_dict() for d in q.stream()]
    enriched = _enrich_attendance(rows)
    if subject_id:
        enriched = [r for r in enriched
                    if r.get("subject_id") == subject_id]
    enriched.sort(key=lambda x: x.get("time", ""))
    return enriched


def _get_subject_id_for_session(row):
    """Helper for filtering by subject_id via session."""
    sid = row.get("session_id")
    if not sid:
        return None
    doc = _col("sessions").document(sid).get()
    return doc.to_dict().get("subject_id") if doc.exists else None


def get_attendance_by_date(query_date, subject_id=None):
    rows = [d.to_dict() for d in
            _col("attendance").where("date", "==", query_date).stream()]
    enriched = _enrich_attendance(rows)
    if subject_id:
        enriched = [r for r in enriched
                    if r.get("subject_id") == subject_id]
    enriched.sort(key=lambda x: x.get("time", ""))
    return enriched


def get_student_history(roll_no):
    docs = (_col("attendance")
            .where("roll_no", "==", roll_no)
            .stream())
    rows = [d.to_dict() for d in docs]
    enriched = _enrich_attendance(rows)
    enriched.sort(key=lambda x: (x.get("date", ""), x.get("time", "")), reverse=True)
    return enriched


def get_attendance_summary(subject_id=None):
    students = get_all_students()
    total = count_sessions(subject_id)

    # Fetch all relevant attendance
    if subject_id:
        session_ids = {s["id"] for s in get_sessions_for_subject(subject_id)}
        att_docs = _col("attendance").stream()
        att_by_roll = {}
        for d in att_docs:
            row = d.to_dict()
            if row.get("session_id") in session_ids:
                att_by_roll[row["roll_no"]] = att_by_roll.get(row["roll_no"], 0) + 1
    else:
        att_docs = _col("attendance").stream()
        att_by_roll = {}
        for d in att_docs:
            rn = d.to_dict()["roll_no"]
            att_by_roll[rn] = att_by_roll.get(rn, 0) + 1

    result = []
    for s in students:
        present = att_by_roll.get(s["roll_no"], 0)
        pct = round((present / total * 100) if total > 0 else 0, 1)
        result.append({
            "roll_no":       s["roll_no"],
            "name":          s["name"],
            "department":    s.get("department", ""),
            "year":          s.get("year", ""),
            "days_present":  present,
            "total_sessions": total,
            "pct":           pct,
        })
    result.sort(key=lambda x: x["days_present"], reverse=True)
    return result


def get_daily_trend(subject_id):
    """Return list of {'date': 'YYYY-MM-DD', 'count': N} for a subject."""
    sessions = get_sessions_for_subject(subject_id)
    session_ids = {s["id"] for s in sessions}
    
    docs = _col("attendance").stream()
    counts_by_date = {}
    
    for d in docs:
        row = d.to_dict()
        if row.get("session_id") in session_ids:
            dt = row.get("date")
            counts_by_date[dt] = counts_by_date.get(dt, 0) + 1
            
    # Sort by date
    dates = sorted(counts_by_date.keys())
    return [{"date": d, "count": counts_by_date[d]} for d in dates]


# ══════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════
def _get_export_rows(from_date=None, to_date=None, subject_id=None):
    docs = _col("attendance").stream()
    rows = [d.to_dict() for d in docs]

    if from_date:
        rows = [r for r in rows if r.get("date", "") >= from_date]
    if to_date:
        rows = [r for r in rows if r.get("date", "") <= to_date]

    enriched = _enrich_attendance(rows)

    if subject_id:
        all_sids = {s["id"] for s in get_sessions_for_subject(subject_id)}
        enriched = [r for r in enriched
                    if any(d.to_dict().get("session_id") in all_sids
                           for d in _col("attendance")
                           .where("roll_no", "==", r["roll_no"])
                           .where("date", "==", r["date"]).stream())]

    enriched.sort(key=lambda x: (x.get("date", ""), x.get("time", "")), reverse=True)
    return enriched


def export_to_csv(output_path=None, from_date=None, to_date=None, subject_id=None):
    os.makedirs(EXPORT_DIR, exist_ok=True)
    if not output_path:
        output_path = os.path.join(EXPORT_DIR, f"attendance_{date.today().isoformat()}.csv")
    rows = _get_export_rows(from_date, to_date, subject_id)
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Roll No","Name","Department","Year",
                         "Subject Code","Subject","Date","Time","Confidence"])
        for r in rows:
            writer.writerow([r["roll_no"], r["name"], r["department"], r["year"],
                             r["subject_code"], r["subject_name"],
                             r["date"], r["time"], r.get("confidence","")])
    return output_path


def export_to_excel(output_path=None, from_date=None, to_date=None, subject_id=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(EXPORT_DIR, exist_ok=True)
    if not output_path:
        output_path = os.path.join(EXPORT_DIR, f"attendance_{date.today().isoformat()}.xlsx")

    rows = _get_export_rows(from_date, to_date, subject_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Roll No","Name","Department","Year",
               "Subject Code","Subject","Date","Time","Confidence"]
    hfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    thin  = Side(style="thin", color="D1D5DB")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center"); c.border = bdr

    for ri, row in enumerate(rows, 2):
        vals = [row["roll_no"], row["name"], row["department"], row["year"],
                row["subject_code"], row["subject_name"],
                row["date"], row["time"], row.get("confidence","")]
        fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") \
               if ri % 2 == 0 else None
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bdr
            if fill: c.fill = fill

    for i, w in enumerate([12,22,14,14,14,24,12,10,12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(output_path)
    return output_path
