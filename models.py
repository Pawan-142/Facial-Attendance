"""
models.py — VisionTrack Database Layer
4-table SQLite schema: students, subjects, sessions, attendance
Replaces the old attendance.py
"""
import sqlite3
import csv
import os
from datetime import datetime, date
from config import DB_FILE, DATABASE_DIR, EXPORT_DIR


# ── Connection helper ──────────────────────────────────────────────────────────
def get_conn():
    os.makedirs(DATABASE_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


# ── Schema init ────────────────────────────────────────────────────────────────
def init_db():
    """Create all tables if they do not exist."""
    conn = get_conn()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS students (
            roll_no     TEXT PRIMARY KEY,
            name        TEXT NOT NULL,
            email       TEXT DEFAULT '',
            department  TEXT DEFAULT '',
            year        TEXT DEFAULT '',
            enrolled_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS subjects (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            code        TEXT NOT NULL UNIQUE,
            name        TEXT NOT NULL,
            department  TEXT DEFAULT '',
            credits     INTEGER DEFAULT 3
        );

        CREATE TABLE IF NOT EXISTS sessions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id  INTEGER NOT NULL,
            date        TEXT NOT NULL,
            start_time  TEXT NOT NULL,
            faculty     TEXT DEFAULT '',
            notes       TEXT DEFAULT '',
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        );

        CREATE TABLE IF NOT EXISTS attendance (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            roll_no     TEXT NOT NULL,
            session_id  INTEGER NOT NULL,
            date        TEXT NOT NULL,
            time        TEXT NOT NULL,
            confidence  REAL,
            UNIQUE(roll_no, session_id),
            FOREIGN KEY (roll_no)     REFERENCES students(roll_no),
            FOREIGN KEY (session_id)  REFERENCES sessions(id)
        );
    """)
    conn.commit()
    conn.close()


# ══════════════════════════════════════════════════════════════
# STUDENTS
# ══════════════════════════════════════════════════════════════
def add_student(roll_no, name, email="", department="", year=""):
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO students (roll_no, name, email, department, year, enrolled_at) "
            "VALUES (?,?,?,?,?,?)",
            (roll_no, name, email, department, year, datetime.now().isoformat())
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()


def update_student(roll_no, name, email="", department="", year=""):
    conn = get_conn()
    conn.execute(
        "UPDATE students SET name=?, email=?, department=?, year=? WHERE roll_no=?",
        (name, email, department, year, roll_no)
    )
    conn.commit()
    conn.close()


def get_all_students():
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM students ORDER BY department, roll_no"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_student(roll_no):
    conn = get_conn()
    row = conn.execute(
        "SELECT * FROM students WHERE roll_no=?", (roll_no,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_student_record(roll_no):
    conn = get_conn()
    conn.execute("DELETE FROM attendance WHERE roll_no=?", (roll_no,))
    conn.execute("DELETE FROM students WHERE roll_no=?", (roll_no,))
    conn.commit()
    conn.close()


# ══════════════════════════════════════════════════════════════
# SUBJECTS
# ══════════════════════════════════════════════════════════════
def add_subject(code, name, department="", credits=3):
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO subjects (code, name, department, credits) VALUES (?,?,?,?)",
            (code.upper(), name, department, credits)
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()


def get_all_subjects():
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM subjects ORDER BY department, code"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_subject(subject_id):
    conn = get_conn()
    row = conn.execute(
        "SELECT * FROM subjects WHERE id=?", (subject_id,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_subject(subject_id):
    conn = get_conn()
    conn.execute(
        "DELETE FROM attendance WHERE session_id IN "
        "(SELECT id FROM sessions WHERE subject_id=?)", (subject_id,)
    )
    conn.execute("DELETE FROM sessions WHERE subject_id=?", (subject_id,))
    conn.execute("DELETE FROM subjects WHERE id=?", (subject_id,))
    conn.commit()
    conn.close()


# ══════════════════════════════════════════════════════════════
# SESSIONS
# ══════════════════════════════════════════════════════════════
def create_session(subject_id, faculty="", notes=""):
    today = date.today().isoformat()
    now   = datetime.now().strftime("%H:%M:%S")
    conn  = get_conn()
    cur   = conn.execute(
        "INSERT INTO sessions (subject_id, date, start_time, faculty, notes) "
        "VALUES (?,?,?,?,?)",
        (subject_id, today, now, faculty, notes)
    )
    session_id = cur.lastrowid
    conn.commit()
    conn.close()
    return session_id


def get_all_sessions():
    conn = get_conn()
    rows = conn.execute("""
        SELECT s.*, sub.code, sub.name AS subject_name
        FROM sessions s
        JOIN subjects sub ON s.subject_id = sub.id
        ORDER BY s.date DESC, s.start_time DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_sessions_for_subject(subject_id):
    conn = get_conn()
    rows = conn.execute("""
        SELECT s.*, sub.code, sub.name AS subject_name
        FROM sessions s
        JOIN subjects sub ON s.subject_id = sub.id
        WHERE s.subject_id=?
        ORDER BY s.date DESC, s.start_time DESC
    """, (subject_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def count_sessions(subject_id=None):
    conn = get_conn()
    if subject_id:
        n = conn.execute(
            "SELECT COUNT(*) FROM sessions WHERE subject_id=?", (subject_id,)
        ).fetchone()[0]
    else:
        n = conn.execute("SELECT COUNT(*) FROM sessions").fetchone()[0]
    conn.close()
    return n


# ══════════════════════════════════════════════════════════════
# ATTENDANCE
# ══════════════════════════════════════════════════════════════
def mark_attendance(roll_no, session_id, confidence=None):
    today = date.today().isoformat()
    now   = datetime.now().strftime("%H:%M:%S")
    conn  = get_conn()
    try:
        conn.execute(
            "INSERT INTO attendance (roll_no, session_id, date, time, confidence) "
            "VALUES (?,?,?,?,?)",
            (roll_no, session_id, today, now, confidence)
        )
        conn.commit()
        return "marked"
    except sqlite3.IntegrityError:
        return "duplicate"
    except Exception as e:
        print(f"[ERROR] mark_attendance: {e}")
        return "error"
    finally:
        conn.close()


def get_session_attendance(session_id):
    conn = get_conn()
    rows = conn.execute("""
        SELECT a.*, st.name, st.department, st.year
        FROM attendance a
        JOIN students st ON a.roll_no = st.roll_no
        WHERE a.session_id=?
        ORDER BY a.time ASC
    """, (session_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_today_attendance(subject_id=None):
    today = date.today().isoformat()
    conn  = get_conn()
    if subject_id:
        rows = conn.execute("""
            SELECT a.roll_no, st.name, st.department, st.year,
                   sub.code, sub.name AS subject_name,
                   a.date, a.time, a.confidence
            FROM attendance a
            JOIN students st  ON a.roll_no    = st.roll_no
            JOIN sessions ses ON a.session_id = ses.id
            JOIN subjects sub ON ses.subject_id = sub.id
            WHERE a.date=? AND ses.subject_id=?
            ORDER BY a.time ASC
        """, (today, subject_id)).fetchall()
    else:
        rows = conn.execute("""
            SELECT a.roll_no, st.name, st.department, st.year,
                   sub.code, sub.name AS subject_name,
                   a.date, a.time, a.confidence
            FROM attendance a
            JOIN students st  ON a.roll_no    = st.roll_no
            JOIN sessions ses ON a.session_id = ses.id
            JOIN subjects sub ON ses.subject_id = sub.id
            WHERE a.date=?
            ORDER BY a.time ASC
        """, (today,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_attendance_by_date(query_date, subject_id=None):
    conn = get_conn()
    if subject_id:
        rows = conn.execute("""
            SELECT a.roll_no, st.name, st.department, st.year,
                   sub.code, sub.name AS subject_name,
                   a.date, a.time, a.confidence
            FROM attendance a
            JOIN students st  ON a.roll_no    = st.roll_no
            JOIN sessions ses ON a.session_id = ses.id
            JOIN subjects sub ON ses.subject_id = sub.id
            WHERE a.date=? AND ses.subject_id=?
            ORDER BY a.time ASC
        """, (query_date, subject_id)).fetchall()
    else:
        rows = conn.execute("""
            SELECT a.roll_no, st.name, st.department, st.year,
                   sub.code, sub.name AS subject_name,
                   a.date, a.time, a.confidence
            FROM attendance a
            JOIN students st  ON a.roll_no    = st.roll_no
            JOIN sessions ses ON a.session_id = ses.id
            JOIN subjects sub ON ses.subject_id = sub.id
            WHERE a.date=?
            ORDER BY a.time ASC
        """, (query_date,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_student_history(roll_no):
    conn = get_conn()
    rows = conn.execute("""
        SELECT a.date, a.time, a.confidence,
               sub.code AS subject_code, sub.name AS subject_name
        FROM attendance a
        JOIN sessions ses ON a.session_id = ses.id
        JOIN subjects sub ON ses.subject_id = sub.id
        WHERE a.roll_no=?
        ORDER BY a.date DESC, a.time DESC
    """, (roll_no,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_attendance_summary(subject_id=None):
    """Per-student attendance stats with percentage."""
    conn = get_conn()
    total = count_sessions(subject_id)

    if subject_id:
        rows = conn.execute("""
            SELECT st.roll_no, st.name, st.department, st.year,
                   COUNT(a.id) AS days_present
            FROM students st
            LEFT JOIN attendance a  ON st.roll_no = a.roll_no
            LEFT JOIN sessions ses  ON a.session_id = ses.id
                                   AND ses.subject_id = ?
            GROUP BY st.roll_no
            ORDER BY days_present DESC
        """, (subject_id,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT st.roll_no, st.name, st.department, st.year,
                   COUNT(a.id) AS days_present
            FROM students st
            LEFT JOIN attendance a ON st.roll_no = a.roll_no
            GROUP BY st.roll_no
            ORDER BY days_present DESC
        """).fetchall()

    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d["total_sessions"] = total
        d["pct"] = round((d["days_present"] / total * 100) if total > 0 else 0, 1)
        result.append(d)
    result.sort(key=lambda x: x["days_present"], reverse=True)
    return result


def get_daily_trend(subject_id):
    """Fallback trend for SQLite."""
    conn = get_conn()
    q = """
        SELECT date, COUNT(DISTINCT roll_no) as count
        FROM attendance a
        JOIN sessions s ON a.session_id = s.id
        WHERE s.subject_id = ?
        GROUP BY date ORDER BY date
    """
    rows = conn.execute(q, (subject_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ══════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════
def _export_query(conn, from_date=None, to_date=None, subject_id=None):
    q = """
        SELECT st.roll_no, st.name, st.department, st.year,
               sub.code AS subject_code, sub.name AS subject_name,
               a.date, a.time, ROUND(a.confidence,4) AS confidence
        FROM attendance a
        JOIN students st  ON a.roll_no    = st.roll_no
        JOIN sessions ses ON a.session_id = ses.id
        JOIN subjects sub ON ses.subject_id = sub.id
        WHERE 1=1
    """
    params = []
    if from_date:  q += " AND a.date >= ?"; params.append(from_date)
    if to_date:    q += " AND a.date <= ?"; params.append(to_date)
    if subject_id: q += " AND ses.subject_id = ?"; params.append(subject_id)
    q += " ORDER BY a.date DESC, a.time ASC"
    return conn.execute(q, params).fetchall()


def export_to_csv(output_path=None, from_date=None, to_date=None, subject_id=None):
    os.makedirs(EXPORT_DIR, exist_ok=True)
    if not output_path:
        output_path = os.path.join(EXPORT_DIR, f"attendance_{date.today().isoformat()}.csv")
    conn = get_conn()
    rows = _export_query(conn, from_date, to_date, subject_id)
    conn.close()
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Roll No", "Name", "Department", "Year",
                         "Subject Code", "Subject", "Date", "Time", "Confidence"])
        writer.writerows(rows)
    return output_path


def export_to_excel(output_path=None, from_date=None, to_date=None, subject_id=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(EXPORT_DIR, exist_ok=True)
    if not output_path:
        output_path = os.path.join(EXPORT_DIR, f"attendance_{date.today().isoformat()}.xlsx")

    conn = get_conn()
    rows = _export_query(conn, from_date, to_date, subject_id)
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Roll No", "Name", "Department", "Year",
               "Subject Code", "Subject", "Date", "Time", "Confidence"]
    hfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)
    thin  = Side(style="thin", color="D1D5DB")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center")
        c.border = bdr

    for ri, row in enumerate(rows, 2):
        fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") \
               if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bdr
            if fill: c.fill = fill

    col_widths = [12, 22, 14, 14, 14, 24, 12, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(output_path)
    return output_path
